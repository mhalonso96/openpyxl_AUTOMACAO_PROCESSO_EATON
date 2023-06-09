import openpyxl

def add_and_check_worksheet(data='', file='', sheetname='', start_row= 1, end_column_add= 5):
    try:
        ##Configuração abertura planilha.
        workbook=  openpyxl.load_workbook(file)
        workbook.active
        worksheet = workbook[sheetname]
        end_row_add = 25065   
        print(f'Ultima linha com valores:', end_row_add)
        values = data
        columns = [column for column in range(self.min_column,self.max_column)]
        #print(columns)
        print('Os dados recebidos são: ', data)
        #print('start row:', start_row, 'end column:',end_column_add)
        for column in columns:
            #print(f'A',column+1)
            for row in range(start_row, end_row_add):
                if worksheet.cell(row=row, column=column+1).value == None:
                    print ('Linha vazia',row)
                    break
                else:
                    print('linha preenchida:', row)
                    row = row
            print(worksheet.cell(row=row, column=column+1).value)
            worksheet.cell(row=row, column=column+1, value = values[column]).number_format = 'Geral'#.number_format ='_-R$ * #.##00_-;-R$ * #.##00_-;_-R$ * "-"_-;_-@_-'
            workbook.save('Planilha.xlsx')
        return print('Salvo com sucesso ..... Linha:', row)

    except Exception as error:
        print(error)
        return error
 