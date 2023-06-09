import openpyxl
def add_and_check_worksheet_currenty(file, data, worksheet, max_row, min_row, max_column, min_column):
    try:    
        workbook=  openpyxl.load_workbook(file)
        workbook.active
        ws = workbook[worksheet]
        last_row = max_row
        #print(f'Ultima linha preenchida', ultima_linha)
        #print(ws.max_column)
        columns = [column for column in range(min_column, max_column+ 1)]
        #print (columns)
        values = data
        print(data)
        for rows in range(min_row, max_row):
            if all ([
                    ws.cell(
                        row= rows, 
                        column=columns[0]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[1]).value == None,
                    ws.cell(
                        row= rows,
                        column=columns[2]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[3]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[4]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[5]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[6]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[7]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[8]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[9]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[10]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[11]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[12]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[13]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[14]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[15]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[16]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[17]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[18]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[19]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[20]).value == None]):
                    print('Linha vazia',rows)
                    row= rows
                    break       
            else:
                last_row +=1
                row= last_row
                print('Linhas Preenchidas', rows)
        ws.cell(
            row= row, 
            column=columns[0], 
            value = values[0])
        ws.cell(
            row= row,
            column=columns[1],
            value = values[1])
        ws.cell(
            row= row,
            column=columns[2], 
            value = values[2])
        ws.cell(
            row= row,
            column=columns[3], 
            value = values[3])
        ws.cell(
            row= row,
            column=columns[4], 
            value = values[4])
        ws.cell(
            row= row,
            column=columns[5], 
            value = values[5])
        ws.cell(
            row= row,
            column=columns[6], 
            value = values[6])
        ws.cell(
            row= row,
            column=columns[7], 
            value = values[7])
        ws.cell(
            row= row,
            column=columns[8], 
            value = values[8])
        ws.cell(
            row= row,
            column=columns[9], 
            value = values[9])
        ws.cell(
            row= row,
            column=columns[10], 
            value = values[10])
        ws.cell(
            row= row,
            column=columns[11], 
            value = values[11])
        ws.cell(
            row= row,
            column=columns[12], 
            value = values[12])
        ws.cell(
            row= row,
            column=columns[13], 
            value = values[13])
        ws.cell(
            row= row,
            column=columns[14], 
            value = values[14])
        ws.cell(
            row= row,
            column=columns[15], 
            value = values[15])
        ws.cell(
            row= row,
            column=columns[16], 
            value = values[16])
        ws.cell(
            row= row,
            column=columns[17], 
            value = values[17])
        ws.cell(
            row= row,
            column=columns[18], 
            value = values[18])
        ws.cell(
            row= row,
            column=columns[19], 
            value = values[19])
        ws.cell(
            row= row, 
            column=columns[20], 
            value = values[20])#.number_format = '_-R$ * #.##00_-;-R$ * #.##00_-;_-R$ * "-"_-;_-@_-'
        workbook.save(file)

    except Exception as error:
        print ('EXCEPTION:', error) 