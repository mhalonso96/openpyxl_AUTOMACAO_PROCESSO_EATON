import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service


class WebScrapingService():
    def __init__(self):
        driver_service = Service(executable_path="chromedriver.exe")
        chrome_options = Options()
        #chrome_options.binary_location = os.getcwd() + os.sep + 'chrome-win' + \
        #    os.sep + 'chrome.exe'
        
        self.driver = webdriver.Chrome(service=driver_service, options=chrome_options)
        

    def start_scrapping(self):
        self.next_page = 1
        self.create_title_sheet()
        self.to_browse_in_site()

    def create_title_sheet(self):
        self.sheet = openpyxl.Workbook()
        self.sheet.create_sheet('Celulares')
        self.phone_sheet = self.sheet['Celulares']
        self.phone_sheet.cell(row=1, column=1, value='Marca')
        self.phone_sheet.cell(row=1, column=2, value='Preco')

        # self.planilha.save('Planilha_teste.xlsx')

    def to_browse_in_site(self):
        self.driver.get('https://telefonesimportados.netlify.app')
        self.info_capture()


    def info_capture(self):
        # Tag xpath para produtos: //h2/a
        # Tag xpath para preos: //div/ins
        produtos = self.driver.find_elements(By.XPATH, '//h2/a')
        precos = self.driver.find_elements(By.XPATH, '//div/ins')
        
        self.save_info(produtos, precos)  
        self.to_browse_in_next_link()      


    def save_info(self, produtos, precos):
        for item in range(0, len(produtos)):
            dados = [
                produtos[item].text,
                precos[item].text,
            ]
            self.phone_sheet.append(dados)

    def to_browse_in_next_link(self):
        self.next_page += 1
        if self.next_page > 5:
            self.driver.quit()
            self.sheet.save('produtos_e-commerce.xlsx')
        else:
            link = f'https://telefonesimportados.netlify.app/shop{self.next_page}.html'
            self.driver.get(link)
            self.info_capture()   