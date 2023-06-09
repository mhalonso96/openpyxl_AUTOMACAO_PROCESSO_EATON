from openpyxl import load_workbook, Workbook

# from utils.add_and_check_worksheet_currenty import add_and_check_worksheet_currenty
class OpenpyxlService(object):
    def __init__(self, wb, filename, data, worksheet, max_row, min_row, max_column, min_column) :
        self.wb = wb
        self. filename = filename
        self.data = data
        self.worksheet = worksheet
        self.max_row = max_row
        self.min_row = min_row
        self.max_column = max_column
        self.min_column = min_column
             
    def get_data(self):
         return self.data

    def get_file(self):
         return self.filename
    
    def get_last_row(self,row):
        return row
    
    def _open_workbook(self):
         self.wb = load_workbook(f'{self.filename}.xlsx')
    
    def save_workbook(self,ws,row,column):
        if not self._cell_verification(ws,row,column):
            self.wb.save(f'{self.filename}.xlsx')
            print('Salvo com sucesso')
        

    def _open_sheet(self):
        self._open_workbook()
        list_sheetnames=self.wb.sheetnames
        for sheetname in list_sheetnames:
            if sheetname == self.worksheet:
                return self.wb[sheetname]
        raise Exception( f'{self.worksheet} não existe na planilha, verique!!')
            
                 
    def controller_sheet(self):

        ws = self._open_sheet()
        #print(ws)
        last_row = self.max_row
        columns = [column for column in range(self.min_column, self.max_column + 1)]
        values = self.data
        ##print ('Colunas: ',columns)
        ##print('data:',self.data, 'min_row:',self.min_row ,'min_row:',self.max_row,)
        
        for rows in range(self.min_row, self.max_row):
            if all ([
                    ws.cell(
                        row= rows, 
                        column=columns[0]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[1]).value == None,
                    ws.cell(
                        row= rows,
                        column=columns[2]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[3]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[4]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[5]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[6]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[7]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[8]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[9]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[10]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[11]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[12]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[13]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[14]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[15]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[16]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[17]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[18]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[19]).value == None,
                    ws.cell(
                        row= rows, 
                        column=columns[20]).value == None]):
                    #print('Linha vazia',rows,'valeu: ', ws.cell(row=row, column=columns[4]).value)
                    row = rows
                    break      
            else:
                last_row += 1
                row= last_row
                #print('Linhas Preenchidas', rows, 'value: ',ws.cell(row=row, column=columns[4]).value)
        ws.cell(
            row= row, 
            column=columns[0], 
            value = values[0])
        ws.cell(
            row= row,
            column=columns[1],
            value = values[1])
        ws.cell(
            row= row,
            column=columns[2], 
            value = values[2])
        ws.cell(
            row= row,
            column=columns[3], 
            value = values[3])
        ws.cell(
            row= row,
            column=columns[4], 
            value = values[4])
        ws.cell(
            row= row,
            column=columns[5], 
            value = values[5])
        ws.cell(
            row= row,
            column=columns[6], 
            value = values[6])
        ws.cell(
            row= row,
            column=columns[7], 
            value = values[7])
        ws.cell(
            row= row,
            column=columns[8], 
            value = values[8])
        ws.cell(
            row= row,
            column=columns[9], 
            value = values[9])
        ws.cell(
            row= row,
            column=columns[10], 
            value = values[10])
        ws.cell(
            row= row,
            column=columns[11], 
            value = values[11])
        ws.cell(
            row= row,
            column=columns[12], 
            value = values[12])
        ws.cell(
            row= row,
            column=columns[13], 
            value = values[13])
        ws.cell(
            row= row,
            column=columns[14], 
            value = values[14])
        ws.cell(
            row= row,
            column=columns[15], 
            value = values[15])
        ws.cell(
            row= row,
            column=columns[16], 
            value = values[16])
        ws.cell(
            row= row,
            column=columns[17], 
            value = values[17])
        ws.cell(
            row= row,
            column=columns[18], 
            value = values[18])
        ws.cell(
            row= row,
            column=columns[19], 
            value = values[19])
        ws.cell(
            row= row, 
            column=columns[20], 
            value = values[20])
        self.save_workbook()
        return print(row)
        
    def _cell_verification(self,ws,row, column):
         
        if ws.cell(row=row, column=column+1).value == None:
            print('check linha',row,'check coluna',column,ws.cell(row=row, column=column).value)
            return True
        else:
            return False


    def controller_sheet_v2(self):
        ws = self._open_sheet()
        values = self.data
    
        for row in range(self.min_row,self.max_row+1):
            for column in range(self.min_column, self.max_column):
                if self._cell_verification(ws,row,column):
                        print('CELULA VAZIA #linha: ',row ,'---- column: ',column ,'----- valor: ', ws.cell(row=row, column=column+1, value=values[column]).value)   
                        
                        
                else:
                    print('CELULA COMPLETA #linha: ',row)
                    break
        

        self.save_workbook(ws,row,column) 
        return row
        
                           
            

        
                
             


    